import os
import openpyxl
import pymysql
import logging

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# 数据库连接信息
db_config = {
    "host": "192.168.3.89",
    "user": "root",
    "password": "123456",
    "database": "jeecg-boot",
    "charset": "utf8mb4"
}

# 连接数据库
def connect_db():
    logging.info("正在连接数据库...")
    connection = pymysql.connect(**db_config)
    logging.info("数据库连接成功。")
    return connection

# 检查单元格是否为合并单元格，并返回实际值
def get_merged_cell_value(sheet, row, col):
    cell = sheet.cell(row=row, column=col)
    for merged_range in sheet.merged_cells.ranges:
        min_row, min_col, max_row, max_col = merged_range.bounds
        if min_row <= row <= max_row and min_col <= col <= max_col:
            if (row == min_row and col == min_col):  # 只读取合并单元格的起始位置
                value = sheet.cell(min_row, min_col).value
                logging.info(f"找到合并单元格: {min_row},{min_col} -> {max_row},{max_col} 值为: {value}")
                return value
            else:
                return None  # 如果不是起始单元格，则返回 None
    value = cell.value
    logging.info(f"单元格 {row},{col} 值为: {value}")
    return value

# 获取培训档案记录
def get_training_records(sheet, start_row, end_row, id_card_number):
    training_records = []
    current_record = {}

    for row in range(start_row, end_row):
        training_type = get_merged_cell_value(sheet, row, 2)
        training_name = get_merged_cell_value(sheet, row, 3)
        training_content = get_merged_cell_value(sheet, row, 4)
        training_time = get_merged_cell_value(sheet, row, 7)
        training_company = get_merged_cell_value(sheet, row, 8)
        training_place = get_merged_cell_value(sheet, row, 9)
        training_teacher = get_merged_cell_value(sheet, row, 10)
        training_hour = get_merged_cell_value(sheet, row, 11)
        training_score = get_merged_cell_value(sheet, row, 12)

        # 处理纵向合并单元格
        if training_type or training_name or training_company:  # 记录的开始
            if current_record:
                training_records.append(current_record)
                current_record = {}  # 重置当前记录

            current_record = {
                "id_card_number": id_card_number,
                "training_type": training_type,
                "training_name": training_name,
                "training_content": training_content,
                "training_time": training_time,
                "training_company": training_company,
                "training_place": training_place,
                "training_teacher": training_teacher,
                "training_hour": training_hour,
                "training_score": training_score,
                "source": "批量导入"
            }
        else:  # 记录的内容拼接
            if current_record:
                if training_name:  # 仅当当前行有培训名称时才拼接
                    current_record["training_name"] += "\n" + training_name
                if training_content:  # 拼接内容
                    current_record["training_content"] += "\n" + training_content

    if current_record:  # 添加最后一条记录
        training_records.append(current_record)

    return training_records

# 根据“锚点”位置读取多条数据
def read_excel_data(file_path):
    logging.info(f"正在读取文件: {file_path}")
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active  # 使用第一个sheet

    # 获取身份证号
    id_card_number = sheet.cell(5, 2).value
    logging.info(f"身份证号: {id_card_number}")

    # 查找证书信息锚点单元格
    certificate_anchor_cell = None
    for row in sheet.iter_rows(min_row=14, max_row=100, min_col=1, max_col=10):
        for cell in row:
            value = get_merged_cell_value(sheet, cell.row, cell.column)
            if value is not None and str(value).strip() == "获取证书信息（执照、职业资格类）":
                certificate_anchor_cell = cell
                logging.info(f"找到证书信息锚点单元格: {cell.coordinate} 值为: {value}")
                break
        if certificate_anchor_cell:
            break

    if not certificate_anchor_cell:
        logging.warning(f"未找到证书信息锚点单元格: {file_path}")
        return [], []

    # 查找培训档案锚点单元格
    training_record_anchor_cell = None
    for row in sheet.iter_rows(min_row=18, max_row=100, min_col=1, max_col=10):
        for cell in row:
            value = get_merged_cell_value(sheet, cell.row, cell.column)
            if value is not None and str(value).strip() == "培训档案":
                training_record_anchor_cell = cell
                logging.info(f"找到培训档案锚点单元格: {cell.coordinate} 值为: {value}")
                break
        if training_record_anchor_cell:
            break

    if not training_record_anchor_cell:
        logging.warning(f"未找到培训档案锚点单元格: {file_path}")
        return [], []

    # 查找诚信档案锚点单元格
    good_faith_record_anchor_cell = None
    for row in sheet.iter_rows(min_row=18, max_row=100, min_col=1, max_col=10):
        for cell in row:
            value = get_merged_cell_value(sheet, cell.row, cell.column)
            if value is not None and str(value).strip() == "诚信档案":
                good_faith_record_anchor_cell = cell
                logging.info(f"找到诚信档案锚点单元格: {cell.coordinate} 值为: {value}")
                break
        if good_faith_record_anchor_cell:
            break

    if not good_faith_record_anchor_cell:
        logging.warning(f"未找到诚信档案锚点单元格: {file_path}")
        return [], []

    # 证书信息数组
    certificate_data_list = []

    # 遍历锚点证书信息下方的多条记录
    certificate_start_row = certificate_anchor_cell.row + 2
    certificate_end_row = training_record_anchor_cell.row
    logging.info(f"证书信息记录范围: {certificate_start_row} 到 {certificate_end_row}")
    for row in range(certificate_start_row, certificate_end_row):
        certificate_type = get_merged_cell_value(sheet, row, 1)
        certificate_name = get_merged_cell_value(sheet, row, 2)
        certificate_company = get_merged_cell_value(sheet, row, 6)
        validity_enddate = get_merged_cell_value(sheet, row, 8)
        remark = get_merged_cell_value(sheet, row, 11)

        # 打印当前记录信息
        logging.info(f"读取证书信息: 类型={certificate_type}, 名称={certificate_name}, 公司={certificate_company}, 结束日期={validity_enddate}, 备注={remark}")

        # 如果数据为空，则结束遍历
        if not certificate_type and not certificate_name and not certificate_company and not validity_enddate and not remark:
            logging.info("已到达证书信息的末尾，停止读取。")
            break

        # 收集数据
        certificate_data_list.append({
            "id_card_number": id_card_number,
            "certificate_type": certificate_type,
            "certificate_name": certificate_name,
            "certificate_company": certificate_company,
            "validity_enddate": validity_enddate,
            "remark": remark
        })

    logging.info(f"读取到的证书信息总数: {len(certificate_data_list)}")

    # 获取培训档案记录
    training_record_start_row = training_record_anchor_cell.row + 2
    training_record_end_row = good_faith_record_anchor_cell.row
    logging.info(f"培训档案范围: {training_record_start_row} 到 {training_record_end_row}")
    training_record_data_list = get_training_records(sheet, training_record_start_row, training_record_end_row, id_card_number)

    logging.info(f"读取到的培训档案信息总数: {len(training_record_data_list)}")
    return certificate_data_list, training_record_data_list

# 将多条数据插入数据库
def insert_data_to_db(data_list, connection, type):
    try:
        if type == "certificate":
            with connection.cursor() as cursor:
                sql = """
                    INSERT INTO training_certificate (id_card_number, certificate_type, certificate_name,
                    certificate_company, validity_enddate, remark) 
                    VALUES (%s, %s, %s, %s, %s, %s)
                """
                for data in data_list:
                    cursor.execute(sql, (
                        data["id_card_number"],
                        data["certificate_type"],
                        data["certificate_name"],
                        data["certificate_company"],
                        data["validity_enddate"],
                        data["remark"]
                    ))
                connection.commit()
                logging.info("证书数据插入成功。")
        elif type == "training":
            with connection.cursor() as cursor:
                sql = """
                    INSERT INTO training_record (id_card_number, training_type, training_name,
                    training_content, training_time, training_company,training_place,
                    training_teacher,training_hour,training_score, source) 
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                for data in data_list:
                    cursor.execute(sql, (
                        data["id_card_number"],
                        data["training_type"],
                        data["training_name"],
                        data["training_content"],
                        data["training_time"],
                        data["training_company"],
                        data["training_place"],
                        data["training_teacher"],
                        data["training_hour"],
                        data["training_score"],
                        data["source"]
                    ))
                connection.commit()
                logging.info("证书数据插入成功。")
    except Exception as e:
        logging.error(f"数据库插入失败: {e}")
        connection.rollback()

# 遍历文件夹，处理所有Excel文件
def process_folder(folder_path):
    logging.info(f"正在处理文件夹: {folder_path}")
    connection = connect_db()
    try:
        for file_name in os.listdir(folder_path):
            if file_name.endswith(".xlsx"):
                file_path = os.path.join(folder_path, file_name)
                certificate_data_list, training_record_data_list = read_excel_data(file_path)
                if certificate_data_list:
                    insert_data_to_db(certificate_data_list, connection, "certificate")
                if training_record_data_list:
                    insert_data_to_db(training_record_data_list, connection, "training")
                    logging.info(f"成功处理文件: {file_name}")
                else:
                    logging.warning(f"文件 {file_name} 没有读取到有效数据。")
    finally:
        connection.close()
        logging.info("数据库连接已关闭。")

folder_path = "C:\\Users\\jingy\\Desktop\\python需求"  # 存放入库信息Excel的文件夹
process_folder(folder_path)
